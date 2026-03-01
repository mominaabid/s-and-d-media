"""
S&D Media – API Views (v2)
Updated for video file uploads instead of URLs
"""
import json
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required
from .models import (
    HeroSection, TrustedLogo,
    PortfolioCategory, PortfolioVideo,
    VideoTestimonial,
    PricingCard, VideoType, Lead,
    TextReview, TeamMember,
)


def json_response(data, status=200):
    return JsonResponse({'success': True, 'data': data}, status=status)

def error_response(message, status=400):
    return JsonResponse({'success': False, 'error': message}, status=status)


class HeroAPIView(View):
    def get(self, request):
        hero = HeroSection.objects.filter(is_active=True).first()
        if not hero:
            return error_response("No active hero found", 404)
        return json_response({
            'heading': hero.heading,
            'subtitle': hero.subtitle,
            'background_video_url': hero.background_video_url or '',
            'background_video_file': request.build_absolute_uri(hero.background_video_file.url) if hero.background_video_file else '',
        })


class LogoBarAPIView(View):
    def get(self, request):
        logos = TrustedLogo.objects.filter(is_active=True)
        return json_response([{
            'id': l.id, 'company_name': l.company_name,
            'logo_url': request.build_absolute_uri(l.logo.url),
            'link': l.link or '',
        } for l in logos])


class PortfolioCategoryAPIView(View):
    def get(self, request):
        cats = PortfolioCategory.objects.filter(is_active=True)
        return json_response([{'id': c.id, 'name': c.name, 'slug': c.slug} for c in cats])


class PortfolioVideoAPIView(View):
    def get(self, request):
        qs = PortfolioVideo.objects.filter(is_active=True)
        if request.GET.get('category'):
            qs = qs.filter(category_id=request.GET['category'])
        return json_response([{
            'id': v.id, 'title': v.title,
            'category': v.category.name, 'category_id': v.category_id,
            'video_file': request.build_absolute_uri(v.video_file.url),
            'thumbnail': request.build_absolute_uri(v.thumbnail.url) if v.thumbnail else '',
            'description': v.description,
        } for v in qs])


class TestimonialAPIView(View):
    def get(self, request):
        items = VideoTestimonial.objects.filter(is_active=True)
        return json_response([{
            'id': t.id, 'name': t.name, 'designation': t.designation,
            'video_file': request.build_absolute_uri(t.video_file.url),
            'thumbnail': request.build_absolute_uri(t.thumbnail.url) if t.thumbnail else '',
        } for t in items])


class PricingAPIView(View):
    def get(self, request):
        cards = PricingCard.objects.filter(is_active=True).prefetch_related('features')
        return json_response([{
            'id': c.id, 'card_type': c.card_type,
            'heading': c.heading, 'subheading': c.subheading,
            'button_label': c.button_label, 'is_most_popular': c.is_most_popular,
            'features': [f.text for f in c.features.all()],
        } for c in cards])


class VideoTypeAPIView(View):
    def get(self, request):
        types = VideoType.objects.filter(is_active=True)
        return json_response([{
            'id': t.id, 'name': t.name,
            'price': float(t.price) if t.price is not None else None
        } for t in types])


@method_decorator(csrf_exempt, name='dispatch')
class LeadSubmitAPIView(View):
    def post(self, request):
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return error_response("Invalid JSON body")

        for field in ['name', 'phone', 'email', 'selected_plan']:
            if not body.get(field):
                return error_response(f"Field '{field}' is required")

        if body['selected_plan'] not in ['bundle', 'custom', 'dedicated']:
            return error_response("selected_plan must be: bundle, custom, or dedicated")

        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
        if ',' in ip:
            ip = ip.split(',')[0].strip()

        lead = Lead.objects.create(
            name=body['name'], company_name=body.get('company_name', ''),
            phone=body['phone'], email=body['email'],
            selected_plan=body['selected_plan'], ip_address=ip or None,
        )

        if body['selected_plan'] == 'custom' and body.get('video_type_ids'):
            lead.selected_video_types.set(
                VideoType.objects.filter(id__in=body['video_type_ids'], is_active=True)
            )

        return json_response({'lead_id': lead.id, 'message': 'Submission received!'}, status=201)


class TextReviewAPIView(View):
    def get(self, request):
        reviews = TextReview.objects.filter(is_active=True)
        return json_response([{
            'id': r.id, 'name': r.name, 'title': r.title,
            'picture': request.build_absolute_uri(r.picture.url) if r.picture else '',
            'review_text': r.review_text, 'rating': r.rating,
        } for r in reviews])


class TeamMemberAPIView(View):
    def get(self, request):
        members = TeamMember.objects.filter(is_active=True)
        return json_response([{
            'id': m.id, 'name': m.name, 'role': m.role, 'bio': m.bio,
            'picture': request.build_absolute_uri(m.picture.url),
            'linkedin_url': m.linkedin_url or '',
        } for m in members])


@staff_member_required(login_url='/cadmin/login/')
def export_leads_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S&D Media Leads"

    hfill = PatternFill(start_color="CDFF00", end_color="CDFF00", fill_type="solid")
    hfont = Font(bold=True, color="222120", size=11)

    headers = ['#', 'Name', 'Company', 'Phone', 'Email', 'Plan', 'Video Types', 'Submitted At', 'IP Address', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    leads = Lead.objects.prefetch_related('selected_video_types').all()
    for row, lead in enumerate(leads, 2):
        video_types = ', '.join(vt.name for vt in lead.selected_video_types.all())
        ws.append([
            row - 1, lead.name, lead.company_name, lead.phone, lead.email,
            lead.get_selected_plan_display(), video_types,
            lead.submitted_at.strftime('%Y-%m-%d %H:%M'),
            lead.ip_address or '', lead.notes
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sd_media_leads.xlsx"'
    wb.save(response)
    return response
