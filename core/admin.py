"""
S&D Media – Admin Panel Configuration
Professional Django admin with:
 - Custom branding (dark grey + neon green)
 - Inline editing where needed
 - Pagination
 - Search & filters
 - Export to Excel
 - Stats dashboard
"""
from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse, path
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models import Count
from django.utils.safestring import mark_safe
from django.contrib import messages
from .models import (
    HeroSection, TrustedLogo,
    PortfolioCategory, PortfolioVideo,
    VideoTestimonial,
    PricingCard, PricingFeature, VideoType, Lead,
    TextReview, TeamMember,
)
import csv
import io


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def image_preview(obj, field='picture', width=60):
    img = getattr(obj, field, None)
    if img:
        return format_html('<img src="{}" width="{}" style="border-radius:6px;object-fit:cover;height:{}px" />', img.url, width, width)
    return format_html('<span style="color:#888">No image</span>')


def video_link(url, label="▶ Preview"):
    if url:
        return format_html('<a href="{}" target="_blank" style="color:#CDFF00">{}</a>', url, label)
    return "–"


def status_badge(active):
    if active:
        return format_html('<span style="background:#CDFF00;color:#222;padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700">ACTIVE</span>')
    return format_html('<span style="background:#555;color:#aaa;padding:2px 10px;border-radius:20px;font-size:11px">INACTIVE</span>')


# ─────────────────────────────────────────────────────────────
# 1. HERO SECTION
# ─────────────────────────────────────────────────────────────
@admin.register(HeroSection)
class HeroSectionAdmin(admin.ModelAdmin):
    list_display = ('heading', 'status_badge_col', 'has_video', 'updated_at')
    list_filter = ('is_active',)
    search_fields = ('heading', 'subtitle')
    readonly_fields = ('updated_at', 'video_preview')
    fieldsets = (
        ('Content', {
            'fields': ('heading', 'subtitle')
        }),
        ('Background Video', {
            'fields': ('background_video_url', 'background_video_file', 'video_preview'),
            'description': 'Provide either a URL (YouTube/Vimeo embed) OR upload a file directly.'
        }),
        ('Settings', {
            'fields': ('is_active', 'updated_at')
        }),
    )

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'
    status_badge_col.allow_tags = True

    def has_video(self, obj):
        if obj.background_video_url or obj.background_video_file:
            return format_html('<span style="color:#CDFF00">✓ Yes</span>')
        return format_html('<span style="color:#888">✗ No</span>')
    has_video.short_description = 'Video?'

    def video_preview(self, obj):
        if obj.background_video_url:
            return video_link(obj.background_video_url, "▶ Open Video URL")
        if obj.background_video_file:
            return video_link(obj.background_video_file.url, "▶ Open Uploaded File")
        return "No video set"
    video_preview.short_description = 'Preview'


# ─────────────────────────────────────────────────────────────
# 2. LOGO BAR
# ─────────────────────────────────────────────────────────────
@admin.register(TrustedLogo)
class TrustedLogoAdmin(admin.ModelAdmin):
    list_display = ('company_name', 'logo_preview', 'link_preview', 'order', 'status_badge_col')
    list_editable = ('order',)
    list_filter = ('is_active',)
    search_fields = ('company_name',)
    list_per_page = 20

    def logo_preview(self, obj):
        return image_preview(obj, 'logo', 50)
    logo_preview.short_description = 'Logo'

    def link_preview(self, obj):
        return video_link(obj.link, obj.link[:40] + '...' if obj.link and len(obj.link) > 40 else obj.link)
    link_preview.short_description = 'Link'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


# ─────────────────────────────────────────────────────────────
# 3. PORTFOLIO
# ─────────────────────────────────────────────────────────────
class PortfolioVideoInline(admin.TabularInline):
    model = PortfolioVideo
    extra = 1
    fields = ('title', 'video_url', 'thumbnail', 'order', 'is_active')
    show_change_link = True


@admin.register(PortfolioCategory)
class PortfolioCategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'slug', 'video_count', 'order', 'status_badge_col')
    list_editable = ('order',)
    prepopulated_fields = {'slug': ('name',)}
    inlines = [PortfolioVideoInline]
    list_per_page = 20

    def video_count(self, obj):
        count = obj.videos.filter(is_active=True).count()
        return format_html('<span style="color:#CDFF00;font-weight:700">{}</span>', count)
    video_count.short_description = 'Active Videos'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


@admin.register(PortfolioVideo)
class PortfolioVideoAdmin(admin.ModelAdmin):
    list_display = ('title', 'category', 'thumbnail_preview', 'video_preview_col', 'order', 'status_badge_col', 'created_at')
    list_filter = ('category', 'is_active')
    search_fields = ('title', 'description')
    list_editable = ('order',)
    list_per_page = 25
    autocomplete_fields = ['category']

    def thumbnail_preview(self, obj):
        return image_preview(obj, 'thumbnail', 55)
    thumbnail_preview.short_description = 'Thumbnail'

    def video_preview_col(self, obj):
        return video_link(obj.video_url)
    video_preview_col.short_description = 'Video'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


# ─────────────────────────────────────────────────────────────
# 4. VIDEO TESTIMONIALS
# ─────────────────────────────────────────────────────────────
@admin.register(VideoTestimonial)
class VideoTestimonialAdmin(admin.ModelAdmin):
    list_display = ('name', 'designation', 'thumbnail_preview', 'video_preview_col', 'order', 'status_badge_col')
    list_editable = ('order',)
    list_filter = ('is_active',)
    search_fields = ('name', 'designation')
    list_per_page = 20

    def thumbnail_preview(self, obj):
        return image_preview(obj, 'thumbnail', 55)
    thumbnail_preview.short_description = 'Thumb'

    def video_preview_col(self, obj):
        return video_link(obj.video_url)
    video_preview_col.short_description = 'Video'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


# ─────────────────────────────────────────────────────────────
# 5. PRICING
# ─────────────────────────────────────────────────────────────
class PricingFeatureInline(admin.TabularInline):
    model = PricingFeature
    extra = 2
    fields = ('text', 'order')
    ordering = ('order',)


@admin.register(PricingCard)
class PricingCardAdmin(admin.ModelAdmin):
    list_display = ('heading', 'card_type_badge', 'is_most_popular', 'button_label', 'order', 'status_badge_col')
    list_editable = ('order',)
    inlines = [PricingFeatureInline]
    list_per_page = 10

    def card_type_badge(self, obj):
        colors = {'bundle': '#4a90d9', 'custom': '#CDFF00', 'dedicated': '#e67e22'}
        color = colors.get(obj.card_type, '#888')
        text_color = '#222' if obj.card_type == 'custom' else '#fff'
        return format_html(
            '<span style="background:{};color:{};padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700">{}</span>',
            color, text_color, obj.get_card_type_display()
        )
    card_type_badge.short_description = 'Type'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


@admin.register(VideoType)
class VideoTypeAdmin(admin.ModelAdmin):
    list_display = ('name', 'price', 'order', 'status_badge_col')
    list_editable = ('price', 'order')
    list_filter = ('is_active',)
    search_fields = ('name',)
    list_per_page = 20

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


# ─────────────────────────────────────────────────────────────
# LEADS – with export + stats
# ─────────────────────────────────────────────────────────────
def export_leads_csv(modeladmin, request, queryset):
    """Admin action: export selected leads to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sd_media_leads.csv"'
    writer = csv.writer(response)
    writer.writerow(['Name', 'Company', 'Phone', 'Email', 'Plan', 'Video Types', 'Submitted At', 'Notes'])
    for lead in queryset:
        video_types = ', '.join(vt.name for vt in lead.selected_video_types.all())
        writer.writerow([
            lead.name, lead.company_name, lead.phone, lead.email,
            lead.get_selected_plan_display(), video_types,
            lead.submitted_at.strftime('%Y-%m-%d %H:%M'), lead.notes
        ])
    return response
export_leads_csv.short_description = "📥 Export selected leads to CSV"


def export_leads_excel_action(modeladmin, request, queryset):
    """Admin action: export selected leads to Excel (.xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Header style
        header_fill = PatternFill(start_color="CDFF00", end_color="CDFF00", fill_type="solid")
        header_font = Font(bold=True, color="222120")

        headers = ['Name', 'Company', 'Phone', 'Email', 'Plan', 'Video Types', 'Submitted At', 'IP', 'Notes']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, lead in enumerate(queryset, 2):
            video_types = ', '.join(vt.name for vt in lead.selected_video_types.all())
            ws.append([
                lead.name, lead.company_name, lead.phone, lead.email,
                lead.get_selected_plan_display(), video_types,
                lead.submitted_at.strftime('%Y-%m-%d %H:%M'),
                lead.ip_address or '', lead.notes
            ])

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sd_media_leads.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        modeladmin.message_user(request, "openpyxl not installed. Run: pip install openpyxl", level=messages.ERROR)
export_leads_excel_action.short_description = "📊 Export selected leads to Excel"


@admin.register(Lead)
class LeadAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'company_name', 'phone', 'email_display',
        'plan_badge', 'video_types_display', 'submitted_at'
    )
    list_filter = ('selected_plan', 'submitted_at')
    search_fields = ('name', 'company_name', 'email', 'phone')
    readonly_fields = ('submitted_at', 'ip_address', 'video_types_display')
    filter_horizontal = ('selected_video_types',)
    actions = [export_leads_csv, export_leads_excel_action]
    list_per_page = 30
    date_hierarchy = 'submitted_at'

    fieldsets = (
        ('Contact Information', {
            'fields': ('name', 'company_name', 'phone', 'email')
        }),
        ('Plan Details', {
            'fields': ('selected_plan', 'selected_video_types')
        }),
        ('Meta', {
            'fields': ('submitted_at', 'ip_address', 'notes'),
            'classes': ('collapse',)
        }),
    )

    def email_display(self, obj):
        return format_html('<a href="mailto:{}" style="color:#CDFF00">{}</a>', obj.email, obj.email)
    email_display.short_description = 'Email'

    def plan_badge(self, obj):
        colors = {'bundle': '#4a90d9', 'custom': '#CDFF00', 'dedicated': '#e67e22'}
        color = colors.get(obj.selected_plan, '#888')
        text_color = '#222' if obj.selected_plan == 'custom' else '#fff'
        return format_html(
            '<span style="background:{};color:{};padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700">{}</span>',
            color, text_color, obj.get_selected_plan_display()
        )
    plan_badge.short_description = 'Plan'

    def video_types_display(self, obj):
        types = obj.selected_video_types.all()
        if types:
            badges = ''.join(
                f'<span style="background:#363432;color:#CDFF00;padding:2px 8px;border-radius:12px;font-size:11px;margin:2px;display:inline-block">{vt.name}</span>'
                for vt in types
            )
            return mark_safe(badges)
        return '–'
    video_types_display.short_description = 'Video Types'


# ─────────────────────────────────────────────────────────────
# 6. TEXT REVIEWS
# ─────────────────────────────────────────────────────────────
@admin.register(TextReview)
class TextReviewAdmin(admin.ModelAdmin):
    list_display = ('name', 'title', 'picture_preview', 'star_rating', 'order', 'status_badge_col')
    list_editable = ('order',)
    list_filter = ('is_active', 'rating')
    search_fields = ('name', 'title', 'review_text')
    list_per_page = 20

    def picture_preview(self, obj):
        return image_preview(obj, 'picture', 50)
    picture_preview.short_description = 'Photo'

    def star_rating(self, obj):
        stars = '★' * obj.rating + '☆' * (5 - obj.rating)
        return format_html('<span style="color:#CDFF00">{}</span>', stars)
    star_rating.short_description = 'Rating'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'


# ─────────────────────────────────────────────────────────────
# 7. TEAM
# ─────────────────────────────────────────────────────────────
@admin.register(TeamMember)
class TeamMemberAdmin(admin.ModelAdmin):
    list_display = ('name', 'role', 'picture_preview', 'linkedin_col', 'order', 'status_badge_col')
    list_editable = ('order',)
    list_filter = ('is_active',)
    search_fields = ('name', 'role')
    list_per_page = 20

    def picture_preview(self, obj):
        return image_preview(obj, 'picture', 55)
    picture_preview.short_description = 'Photo'

    def linkedin_col(self, obj):
        return video_link(obj.linkedin_url, "LinkedIn")
    linkedin_col.short_description = 'LinkedIn'

    def status_badge_col(self, obj):
        return status_badge(obj.is_active)
    status_badge_col.short_description = 'Status'
